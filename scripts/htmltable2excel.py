#!/usr/bin/env python3
"""
htmltable2excel

读取 Markdown 文件中的 HTML 表格，提取所有表格数据，
以第一个表格的表头为准合并为一个 Excel（.xlsx）并写入输出文件。

Usage:
  python scripts/htmltable2excel.py \
    --input data/20251023html-table-dataA1.md \
    --output data/20251023excel-dataA1.xlsx

Notes / 说明:
- 只使用第一个表格的表头（项目需求表示表头一致）。
- 单元格内容保留换行符（在 Excel 中建议开启自动换行查看）。
- 解析依赖 beautifulsoup4；写入 Excel 依赖 openpyxl。
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Iterable, List, Tuple

from bs4 import BeautifulSoup


def parse_tables_from_markdown(md_text: str) -> List[List[List[str]]]:
    """Parse all HTML tables embedded in a Markdown document.

    Returns a list of tables, where each table is a list of rows,
    and each row is a list of cell strings.
    """
    soup = BeautifulSoup(md_text, "html.parser")
    tables = []
    for tbl in soup.find_all("table"):
        rows: List[List[str]] = []
        for tr in tbl.find_all("tr"):
            cells = tr.find_all(["th", "td"])  # header or data cells
            # Normalize cell text: strip surrounding whitespace; preserve inner newlines
            row = [c.get_text(separator="\n", strip=True) for c in cells]
            if row:  # skip empty rows
                rows.append(row)
        if rows:
            tables.append(rows)
    return tables


def merge_tables_with_first_header(tables: List[List[List[str]]]) -> Tuple[List[str], List[List[str]]]:
    """Merge multiple tables using the first table's header.

    - Uses the first row of the first table as the header.
    - Appends all subsequent rows from all tables.
    - If a row length mismatches the header, it is padded/truncated.
    """
    if not tables:
        return [], []

    header = tables[0][0]
    n_cols = len(header)

    merged_rows: List[List[str]] = []
    for table in tables:
        # Skip the first row (header) for every table
        data_rows = table[1:]
        for row in data_rows:
            if len(row) < n_cols:
                # pad with empty strings to match header length
                padded = row + [""] * (n_cols - len(row))
                merged_rows.append(padded)
            elif len(row) > n_cols:
                # truncate extras if any
                merged_rows.append(row[:n_cols])
            else:
                merged_rows.append(row)

    return header, merged_rows


def write_excel(path: Path, header: List[str], rows: Iterable[List[str]]) -> None:
    """Write data to an .xlsx Excel file using openpyxl.

    - Creates directories as needed.
    - Writes header then all rows.
    - Enables wrap_text so multi-line cell content displays well.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment
    except Exception as e:  # pragma: no cover - helpful error for users
        raise RuntimeError(
            "openpyxl is required for Excel output. Install with: pip install openpyxl"
        ) from e

    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    wrap_align = Alignment(wrap_text=True)

    # Write header
    if header:
        ws.append(header)
        for col_idx in range(1, len(header) + 1):
            ws.cell(row=1, column=col_idx).alignment = wrap_align

    # Write data rows
    start_row = 2 if header else 1
    for r_idx, row in enumerate(rows, start=start_row):
        ws.append(row)
        for c_idx in range(1, len(row) + 1):
            ws.cell(row=r_idx, column=c_idx).alignment = wrap_align

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract HTML tables in Markdown to merged Excel (.xlsx).")
    parser.add_argument(
        "--input",
        default="data/20251023html-table-dataA1.md",
        type=str,
        help="Input Markdown file path containing HTML <table> blocks.",
    )
    parser.add_argument(
        "--output",
        default="data/20251023excel-dataA1.xlsx",
        type=str,
        help="Output Excel (.xlsx) file path.",
    )
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    md_text = input_path.read_text(encoding="utf-8")
    tables = parse_tables_from_markdown(md_text)
    if not tables:
        raise ValueError("No <table> found in input Markdown.")

    header, rows = merge_tables_with_first_header(tables)
    if not header:
        raise ValueError("Failed to detect header in the first table.")

    write_excel(output_path, header, rows)

    # Minimal, helpful console output
    print(f"Parsed tables: {len(tables)}")
    print(f"Rows written: {len(rows)} (header columns: {len(header)})")
    print(f"Excel saved to: {output_path}")


if __name__ == "__main__":
    main()

